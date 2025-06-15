import os
import re
import logging
from datetime import datetime
from django.shortcuts import render
from django.http import JsonResponse, HttpResponseServerError
from django.conf import settings
from openpyxl import load_workbook, Workbook
from openpyxl.utils.exceptions import InvalidFileException

# Set up logging
logger = logging.getLogger(__name__)

def is_valid_email(email):
    """Validate email format"""
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return re.match(pattern, email) is not None

def get_excel_path():
    """Get the path to the Excel file, creating directory if needed"""
    try:
        # Create a 'subscriptions' directory in the project root if it doesn't exist
        excel_dir = os.path.join(settings.BASE_DIR, 'subscriptions')
        if not os.path.exists(excel_dir):
            os.makedirs(excel_dir, exist_ok=True)
            logger.info(f"Created directory: {excel_dir}")
            
        # Ensure directory is writable
        if not os.access(excel_dir, os.W_OK):
            raise PermissionError(f"Cannot write to directory: {excel_dir}")
            
        excel_path = os.path.join(excel_dir, 'subscribers.xlsx')
        logger.info(f"Using Excel file: {excel_path}")
        return excel_path
        
    except Exception as e:
        logger.error(f"Error getting Excel path: {str(e)}")
        raise

def save_to_excel(email):
    """Save email to Excel file with error handling"""
    if not is_valid_email(email):
        return False, "Invalid email format"
        
    excel_path = get_excel_path()
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    try:
        # Try to load existing workbook or create new one
        if os.path.exists(excel_path):
            try:
                wb = load_workbook(excel_path)
                ws = wb.active
                # Check if header exists, add if not
                if ws.max_row == 0 or not all(cell.value in ['Email', 'Subscription Date'] for cell in ws[1] if cell.value):
                    ws.append(['Email', 'Subscription Date'])
            except (InvalidFileException, KeyError, IndexError) as e:
                logger.warning(f"Existing Excel file is invalid or corrupted, creating a new one: {str(e)}")
                wb = Workbook()
                ws = wb.active
                ws.title = "Subscribers"
                ws.append(['Email', 'Subscription Date'])
        else:
            # Create new workbook if file doesn't exist
            wb = Workbook()
            ws = wb.active
            ws.title = "Subscribers"
            ws.append(['Email', 'Subscription Date'])
        
        # Check for existing email (case-insensitive)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0] and row[0].lower() == email.lower():
                return False, "This email is already subscribed!"
        
        # Add new subscription
        try:
            ws.append([email, timestamp])
            wb.save(excel_path)
            logger.info(f"Successfully saved subscription for {email}")
            return True, "Thank you for subscribing!"
            
        except PermissionError as e:
            error_msg = f"Permission denied when saving to {excel_path}: {str(e)}"
            logger.error(error_msg)
            return False, "Error saving subscription. Please try again later. If the problem persists, please contact support."
            
    except Exception as e:
        error_msg = f"Unexpected error in save_to_excel: {str(e)}"
        logger.error(error_msg, exc_info=True)
        return False, "An unexpected error occurred. Our team has been notified. Please try again later."

def index(request):
    if request.method == 'POST' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        email = (request.POST.get('email') or '').strip()
        
        if not email:
            return JsonResponse(
                {'status': 'error', 'message': 'Email is required'}, 
                status=400
            )
        
        try:
            success, message = save_to_excel(email)
            if success:
                return JsonResponse(
                    {'status': 'success', 'message': 'Thank you for subscribing!'}
                )
            else:
                return JsonResponse(
                    {'status': 'error', 'message': message},
                    status=400
                )
                
        except Exception as e:
            logger.error(f"Error in index view: {str(e)}")
            return JsonResponse(
                {'status': 'error', 'message': 'An error occurred. Please try again later.'}, 
                status=500
            )
    
    # For GET requests, just render the page
    return render(request, 'index.html')
